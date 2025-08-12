import datetime
from openpyxl import Workbook
from openpyxl.styles import Font


class ClientTraffic:
    def __init__(self, raw_line: str):
        self.raw_line = raw_line.strip()
        self.client_ip = None
        self.traffic_bytes = 0
        self.date_from = None
        self.date_to = None
        self.parse_line()

    def parse_line(self):
        parts = self.raw_line.split()
        if len(parts) != 4:
            print(self.raw_line)
            raise ValueError(f"Неверный формат строки: {self.raw_line}")

        name_part, traffic_str, date_from_str, date_to_str = parts

        if not name_part.startswith("client_") or "_download" not in name_part:
            raise ValueError(f"Неверное имя клиента: {name_part}")

        self.client_ip = name_part.replace("client_", "").replace("_download", "")
        self.traffic_bytes = int(traffic_str)
        self.date_from = datetime.datetime.strptime(date_from_str, "%Y-%m-%d").date()
        self.date_to = datetime.datetime.strptime(date_to_str, "%Y-%m-%d").date()

    def traffic_gb(self):
        return round(self.traffic_bytes / 1_000_000_000, 2)  # GB (10^9)

    def duration_days(self):
        return (self.date_to - self.date_from).days

    def to_row(self):
        return [
            self.client_ip,
            self.traffic_gb(),
            self.date_from.isoformat(),
            self.date_to.isoformat(),
            self.duration_days()
        ]


class TrafficReport:
    def __init__(self, input_file: str, output_file: str):
        self.input_file = input_file
        self.output_file = output_file
        self.clients = []

    def load_data(self):
        with open(self.input_file, "r") as f:
            for line in f:
                stripped = line.strip()
                if stripped and stripped.startswith("client_"):
                    try:
                        client = ClientTraffic(stripped)
                        self.clients.append(client)
                    except ValueError as e:
                        print(f"Пропущена строка из-за ошибки: {e}")

    def generate_excel(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Client Traffic"

        header = ["Client", "Traffic used (GB)", "From", "To", "In days"]
        ws.append(header)

        total_traffic = 0.0

        for client in self.clients:
            row = client.to_row()
            total_traffic += row[1]
            ws.append(row)

        # Пустая строка и итог
        ws.append([])
        summary_row = ["Total", round(total_traffic, 2), "", "", ""]
        ws.append(summary_row)

        # Жирный шрифт для заголовка и итога
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)

        wb.save(self.output_file)


if __name__ == "__main__":
    report = TrafficReport("data.txt", "client_traffic_report.xlsx")
    report.load_data()
    report.generate_excel()
    print("Отчёт успешно создан: client_traffic_report.xlsx")
